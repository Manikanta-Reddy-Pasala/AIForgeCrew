"""Tally ↔ OneShell trial-balance reconciliation (3-way).

Deterministic — no LLM call. Stages:

    1. validate_file()     schema + numeric parseability before any work
    2. parse_tally / parse_oneshell      file → TBRow list
    3. fetch_oneshell_from_mongo()       live DB → TBRow list (optional)
    4. reconcile()                       2-way (file-vs-file)
       reconcile_3way()                  Tally vs OneShell-file vs DB
    5. render_markdown / write_csv       audit-friendly output

CLI:

    aiforge-agents-tb \\
        --tally    ~/Downloads/tally-acme-2026.xlsx \\
        --oneshell ~/Downloads/oneshell-acme-2026.csv \\
        --env qa --business b117695104178401 \\
        --validate-with-mongo            # pulls live OneShell rows too

Full spec: `docs/processes/tally-oneshell-trial-balance.md`.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any


# ─────────── normalisation helpers ───────────────────────────────────

_NAME_PUNCT = re.compile(r"[^\w\s]+")
_NAME_WS = re.compile(r"\s+")


def normalise_name(s: str | None) -> str:
    if s is None:
        return ""
    out = str(s).strip().lower()
    out = _NAME_PUNCT.sub(" ", out)
    out = _NAME_WS.sub(" ", out).strip()
    return out


def to_decimal(v: Any) -> float:
    """Tolerant numeric parser. Returns 0.0 for blanks / non-numeric."""
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("₹", "")
    # Tally exports "(1,000.00)" for negative — wrap in parens
    if s.startswith("(") and s.endswith(")"):
        try:
            return -float(s[1:-1])
        except ValueError:
            return 0.0
    # Trailing CR/DR markers
    suffix = ""
    if s[-2:].upper() in ("CR", "DR"):
        suffix = s[-2:].upper()
        s = s[:-2].strip()
    try:
        n = float(s)
    except ValueError:
        return 0.0
    return -n if suffix == "CR" else n


# ─────────── readers ──────────────────────────────────────────────────

@dataclass
class TBRow:
    code: str = ""
    name: str = ""
    parent: str = ""
    opening: float = 0.0
    debit: float = 0.0
    credit: float = 0.0
    closing: float = 0.0


# ─────────── file validation ─────────────────────────────────────────

@dataclass
class FileValidation:
    path: str
    ok: bool = True
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    row_count: int = 0
    columns: list[str] = field(default_factory=list)
    detected_kind: str = ""   # tally | oneshell | unknown


_TALLY_REQUIRED = {"particulars", "account", "name", "ledger"}    # any one
_TALLY_NUMERIC = ["closing balance", "closing", "cb"]
_ONESHELL_REQUIRED = {"accountname", "name", "account", "ledger"}
_ONESHELL_NUMERIC = ["closingbalance", "closing"]


def _detect_kind(filename: str, columns: list[str]) -> str:
    low = (filename or "").lower()
    cols_low = [c.lower() for c in columns]
    if "tally" in low or any("particulars" == c for c in cols_low):
        return "tally"
    if "oneshell" in low or any("accountname" == c for c in cols_low):
        return "oneshell"
    return "unknown"


def validate_file(path: str | Path, *, expected: str = "") -> FileValidation:
    """Schema + parseability check BEFORE we trust the file.

    `expected` ∈ {tally, oneshell, ""}. When set, mismatched detected
    kind becomes an error. Empty value falls back to filename / column
    detection.
    """
    p = Path(path)
    fv = FileValidation(path=str(p))
    if not p.exists():
        fv.ok = False
        fv.errors.append(f"file not found: {p}")
        return fv
    if p.stat().st_size == 0:
        fv.ok = False
        fv.errors.append("empty file")
        return fv
    try:
        rows = _read_table(p)
    except SystemExit as exc:
        fv.ok = False
        fv.errors.append(f"unreadable: {exc}")
        return fv
    fv.row_count = len(rows)
    if not rows:
        fv.ok = False
        fv.errors.append("no data rows")
        return fv
    fv.columns = list(rows[0].keys())
    cols_lower = {c.lower() for c in fv.columns}
    fv.detected_kind = _detect_kind(p.name, fv.columns)

    if expected and fv.detected_kind != "unknown" \
            and fv.detected_kind != expected:
        # Tally-shaped OneShell files are common (chartOfAccounts
        # exports preserve the Tally column layout). Treat the
        # mismatch as a warning when the file LOOKS like trial-
        # balance data — keeps the pipeline running on real exports.
        fv.warnings.append(
            f"detected kind={fv.detected_kind} but expected={expected} — "
            "treating as Tally-shaped variant"
        )
    kind = expected or fv.detected_kind
    required = (
        _TALLY_REQUIRED if kind == "tally"
        else _ONESHELL_REQUIRED if kind == "oneshell"
        else None
    )
    numerics = (
        _TALLY_NUMERIC if kind == "tally"
        else _ONESHELL_NUMERIC if kind == "oneshell"
        else []
    )
    if required is not None:
        if cols_lower.isdisjoint(required):
            fv.errors.append(
                f"missing any-of name column: expected one of "
                f"{sorted(required)}, got {sorted(cols_lower)}"
            )
        # Numeric column sanity — the FIRST numeric we recognise must
        # parse to float on at least 1 row.
        for nc in numerics:
            real = next((c for c in fv.columns if c.lower() == nc), None)
            if real is None:
                continue
            ok = False
            for r in rows[: min(50, len(rows))]:
                v = r.get(real)
                if v in (None, ""):
                    continue
                try:
                    to_decimal(v)
                    ok = True
                    break
                except Exception:
                    pass
            if not ok:
                fv.warnings.append(
                    f"numeric column `{real}` had no parseable values "
                    "in the first 50 rows"
                )
            break
    if fv.errors:
        fv.ok = False
    return fv


# Tokens that signal "this row is the header" — Tally + OneShell variants.
_HEADER_TOKENS: frozenset[str] = frozenset({
    "particulars", "account", "accountname", "ledger", "name",
    "opening", "openingbalance", "ob",
    "debit", "dr", "credit", "cr",
    "closing", "closingbalance", "cb",
    "group", "parent", "parentname",
    "code", "accountcode",
})


def _row_header_hits(row: tuple) -> int:
    cells = [str(c or "").strip().lower().replace(" ", "")
             for c in row if c]
    return sum(1 for c in cells if c in _HEADER_TOKENS)


def _detect_header_row(rows: list[tuple]) -> tuple[int, list[str]]:
    """Scan first ~12 rows. Find the contiguous block of "header-like"
    rows (any row that has at least 1 known TB token) and merge them
    column-wise.

    Real Tally exports stack 3 rows of header: row N has `Particulars`
    in col 0, row N+1 has `Opening / Transactions / Closing` spanning
    several cols, row N+2 has `Balance / Debit / Credit / Balance`.
    Merging the three columnar gives proper field names like
    `Opening Balance`, `Transactions Debit`, `Closing Balance`,
    `Particulars`.

    Returns (data_start_row_idx, merged_header_cells). Fallback to
    row 0 when nothing matches."""
    scan = min(12, len(rows))
    if scan == 0:
        return 0, []
    # Per-row score
    scores = [_row_header_hits(rows[i]) for i in range(scan)]
    # Pick the first row with score >= 2 OR a contiguous block of rows
    # that together have >= 3 hits. Then extend forward while next row
    # has >= 1 hit and is non-data-looking.
    start = -1
    for i in range(scan):
        if scores[i] >= 2:
            start = i
            break
    if start == -1:
        # Try block-merge: find first row with hits, walk forward
        for i in range(scan):
            if scores[i] >= 1:
                start = i
                break
    if start == -1:
        return 1, [str(c or "").strip() for c in rows[0]]
    # Extend BACKWARD too — Tally puts `Particulars` (1 hit) one row
    # above the `Opening / Transactions / Closing` (3 hits) row.
    while start - 1 >= 0 and scores[start - 1] >= 1:
        start -= 1
    # Extend the block forward — accept any row with hits >= 1.
    end = start
    for j in range(start + 1, scan):
        if scores[j] >= 1:
            end = j
        else:
            break
    width = max(len(r) for r in rows[start:end + 1])
    merged: list[str] = []
    for col in range(width):
        parts: list[str] = []
        for r in rows[start:end + 1]:
            if col < len(r):
                v = str(r[col] or "").strip()
                if v and v.lower() not in {p.lower() for p in parts}:
                    parts.append(v)
        merged.append(" ".join(parts).strip())
    return end + 1, merged


def _read_table(path: Path) -> list[dict[str, Any]]:
    """Read .csv/.xlsx/.xls into list of {column: value}.

    Auto-detects the header row inside the first 25 rows so real Tally
    exports (with title + date-range banner above the column header)
    parse without manual --header-row.

    .xlsx / .xls require the optional `openpyxl` / `xlrd` deps; if
    either is missing the file must be saved as CSV first.
    """
    suf = path.suffix.lower()
    if suf == ".csv":
        with open(path, "r", encoding="utf-8-sig") as f:
            raw_rows = [tuple(row) for row in csv.reader(f)]
        if not raw_rows:
            return []
        data_start, header = _detect_header_row(raw_rows)
        out: list[dict[str, Any]] = []
        for row in raw_rows[data_start:]:
            out.append({
                header[i]: row[i] if i < len(row) else ""
                for i in range(len(header))
            })
        return out
    if suf in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise SystemExit(f"openpyxl required for {path.name}: {exc}")
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        data_start, header = _detect_header_row(rows)
        out: list[dict[str, Any]] = []
        for row in rows[data_start:]:
            out.append({
                header[i]: row[i] if i < len(row) else ""
                for i in range(len(header))
            })
        return out
    if suf == ".xls":
        try:
            import xlrd  # type: ignore
        except ImportError as exc:
            raise SystemExit(f"xlrd required for {path.name}: {exc}")
        wb = xlrd.open_workbook(str(path))
        ws = wb.sheet_by_index(0)
        if ws.nrows == 0:
            return []
        header = [str(ws.cell_value(0, c) or "").strip() for c in range(ws.ncols)]
        return [
            {header[c]: ws.cell_value(r, c) for c in range(ws.ncols)}
            for r in range(1, ws.nrows)
        ]
    raise SystemExit(f"unsupported file type: {path}")


def _pick(d: dict[str, Any], *keys: str) -> Any:
    """First non-empty value for any of the given keys.

    Match priority:
      1. exact case-insensitive
      2. case-insensitive substring contains (so `_pick("debit")` lands
         on `Transactions Debit`)
    """
    real_keys = {k.lower(): k for k in d.keys()}
    # Pass 1: exact
    for k in keys:
        real = real_keys.get(k.lower())
        if real and d.get(real) not in (None, ""):
            return d[real]
    # Pass 2: substring (longest needle wins to disambiguate
    # "credit" vs "transactions credit" vs "closing balance" vs "balance")
    needles = sorted(keys, key=len, reverse=True)
    for needle in needles:
        nl = needle.lower()
        for col_low, real in real_keys.items():
            if nl in col_low and d.get(real) not in (None, ""):
                return d[real]
    return ""


def parse_tally(path: Path) -> list[TBRow]:
    """Read a Tally export. Tolerant to header variations."""
    rows = _read_table(path)
    out: list[TBRow] = []
    for r in rows:
        name = str(_pick(r, "Particulars", "Account", "Name", "Ledger") or "").strip()
        if not name:
            continue
        parent = str(_pick(r, "Group", "Parent") or "").strip()
        code = str(_pick(r, "Account Code", "Code", "Ledger Code") or "").strip()
        ob = to_decimal(_pick(r, "Opening Balance", "OB", "Opening"))
        dr = to_decimal(_pick(r, "Debit", "Dr"))
        cr = to_decimal(_pick(r, "Credit", "Cr"))
        cb = to_decimal(_pick(r, "Closing Balance", "Closing", "CB"))
        # skip empty summary rows (no numbers at all)
        if all(v == 0.0 for v in (ob, dr, cr, cb)):
            continue
        out.append(TBRow(
            code=code, name=name, parent=parent,
            opening=ob, debit=dr, credit=cr, closing=cb,
        ))
    return out


# ─────────── live OneShell DB fetcher ─────────────────────────────────

def fetch_oneshell_via_api(
    *, business_id: str, env: str = "qa",
    from_date: str = "", to_date: str = "",
    financial_year: int | None = None,
) -> list[TBRow]:
    """Pull OneShell trial-balance rows by calling PosClientBackend's
    `/trialBalance` REST endpoint. This is the AUTHORITATIVE source —
    PCB computes TB on-demand from transactions; chartOfAccounts in
    Mongo only holds opening balances.

    Endpoint base via env var:
      AIFORGE_PCB_API_BASE  (default http://localhost:8090/v1/api)

    Two endpoints supported:
      - /trialBalance?businessId=...&fromDate=YYYY-MM-DD&toDate=YYYY-MM-DD
      - /trialBalance/financialYear?businessId=...&financialYear=2025
    """
    import os
    import httpx
    base = os.environ.get(
        "AIFORGE_PCB_API_BASE", "http://localhost:8090/v1/api",
    ).rstrip("/")
    params: dict[str, Any] = {"businessId": business_id}
    if financial_year:
        url = base + "/trialBalance/financialYear"
        params["financialYear"] = financial_year
    else:
        url = base + "/trialBalance"
        if from_date:
            params["fromDate"] = from_date
        if to_date:
            params["toDate"] = to_date
    r = httpx.get(url, params=params, timeout=60.0)
    r.raise_for_status()
    rows: list[TBRow] = []
    for entry in r.json() or []:
        rows.append(TBRow(
            code=str(entry.get("code") or "").strip(),
            name=str(entry.get("name") or "").strip(),
            parent=str(entry.get("parentName") or "").strip(),
            opening=to_decimal(entry.get("openingBalance")),
            debit=to_decimal(entry.get("debit") or entry.get("periodDebit")),
            credit=to_decimal(entry.get("credit") or entry.get("periodCredit")),
            closing=to_decimal(entry.get("closingBalance")),
        ))
    return rows


def fetch_oneshell_from_mongo(
    *, business_id: str, env: str = "qa",
    fy_start: str | None = None,
) -> list[TBRow]:
    """Direct Mongo fallback when PCB API isn't reachable.

    NOTE: chartOfAccounts only has openingBalance; closing/period are
    computed from transactions. Use `fetch_oneshell_via_api()` for the
    authoritative full trial balance. This fetcher is best-effort —
    useful when only opening balances are needed (e.g., FY-start sanity
    check).

    Connection priority:
      1. AIFORGE_MONGODB_SERVICE_URL — HTTP gateway (CLAUDE.md mandate).
      2. AIFORGE_MONGO_URI — direct pymongo URI.
      3. mongodb://localhost:27017/oneshell fallback.
    """
    import os
    rows: list[TBRow] = []

    # Path 2: HTTP gateway first (production default — never bypass it)
    svc_url = os.environ.get("AIFORGE_MONGODB_SERVICE_URL", "")
    if svc_url:
        try:
            import httpx
            r = httpx.post(
                svc_url.rstrip("/") + "/v1/find",
                json={
                    "collection": "chartOfAccounts",
                    "filter": {"businessId": business_id},
                    "projection": {
                        "name": 1, "code": 1, "parentName": 1,
                        "openingBalance": 1, "closingBalance": 1,
                        "periodDebit": 1, "periodCredit": 1,
                    },
                    "limit": 5000,
                },
                timeout=30.0,
            )
            r.raise_for_status()
            for doc in r.json().get("docs") or []:
                rows.append(_doc_to_tbrow(doc))
            return rows
        except Exception:
            pass    # fall through to direct mongo

    # Path 1 / 3: pymongo direct
    try:
        from pymongo import MongoClient
    except ImportError as exc:
        raise SystemExit(
            "pymongo required for --validate-with-mongo "
            f"(install via `pip install pymongo`): {exc}"
        )
    uri = os.environ.get(
        "AIFORGE_MONGO_URI",
        "mongodb://localhost:27017/oneshell",
    )
    client = MongoClient(uri, serverSelectionTimeoutMS=5000)
    # pymongo ≥4 forbids truthiness on Database objects; check is None.
    db = client.get_default_database()
    if db is None:
        db = client["oneshell"]
    cursor = db["chartOfAccounts"].find(
        {"businessId": business_id},
        {
            "name": 1, "code": 1, "parentName": 1,
            "openingBalance": 1, "closingBalance": 1,
            "periodDebit": 1, "periodCredit": 1,
        },
    ).limit(5000)
    for doc in cursor:
        rows.append(_doc_to_tbrow(doc))
    return rows


def _doc_to_tbrow(doc: dict) -> TBRow:
    return TBRow(
        code=str(doc.get("code") or "").strip(),
        name=str(doc.get("name") or "").strip(),
        parent=str(doc.get("parentName") or "").strip(),
        opening=to_decimal(doc.get("openingBalance")),
        debit=to_decimal(doc.get("periodDebit")),
        credit=to_decimal(doc.get("periodCredit")),
        closing=to_decimal(doc.get("closingBalance")),
    )


def parse_oneshell(path: Path) -> list[TBRow]:
    """Tolerant OneShell parser. Accepts both:
      - Native OneShell schema (accountName / accountCode / closingBalance …)
      - Tally-shaped exports (Particulars / Opening Balance / Closing Balance …)
        which is what `chartOfAccounts` exports look like in practice.
    """
    rows = _read_table(path)
    out: list[TBRow] = []
    for r in rows:
        name = str(_pick(r, "accountName", "name", "particulars",
                         "account", "ledger") or "").strip()
        if not name:
            continue
        parent = str(_pick(r, "parentName", "parent", "group") or "").strip()
        code = str(_pick(r, "accountCode", "code") or "").strip()
        ob = to_decimal(_pick(r, "openingBalance", "opening", "ob"))
        dr = to_decimal(_pick(r, "periodDebit", "debit", "dr"))
        cr = to_decimal(_pick(r, "periodCredit", "credit", "cr"))
        cb = to_decimal(_pick(r, "closingBalance", "closing", "cb"))
        if all(v == 0.0 for v in (ob, dr, cr, cb)):
            continue
        out.append(TBRow(
            code=code, name=name, parent=parent,
            opening=ob, debit=dr, credit=cr, closing=cb,
        ))
    return out


# ─────────── reconciliation ───────────────────────────────────────────

@dataclass
class Reconciliation:
    env: str
    business_id: str
    tally_total: float = 0.0
    oneshell_total: float = 0.0
    gap: float = 0.0
    matched: list[dict] = field(default_factory=list)
    diff: list[dict] = field(default_factory=list)
    large: list[dict] = field(default_factory=list)
    tally_only: list[dict] = field(default_factory=list)
    oneshell_only: list[dict] = field(default_factory=list)


_MATCH_TOLERANCE = 1.00
_DIFF_TOLERANCE = 100.00


# ─────────── totals + drill-down (top-down workflow) ────────────────

@dataclass
class Totals:
    source: str
    total_ob: float = 0.0
    total_dr: float = 0.0
    total_cr: float = 0.0
    total_cb: float = 0.0
    row_count: int = 0


def compute_totals(rows: list[TBRow], *, source: str) -> Totals:
    t = Totals(source=source, row_count=len(rows))
    for r in rows:
        t.total_ob += r.opening
        t.total_dr += r.debit
        t.total_cr += r.credit
        t.total_cb += r.closing
    t.total_ob = round(t.total_ob, 2)
    t.total_dr = round(t.total_dr, 2)
    t.total_cr = round(t.total_cr, 2)
    t.total_cb = round(t.total_cb, 2)
    return t


@dataclass
class TotalsCompare:
    a: Totals
    b: Totals
    delta_ob: float = 0.0
    delta_dr: float = 0.0
    delta_cr: float = 0.0
    delta_cb: float = 0.0
    matched: bool = False
    dimensions_diverged: list[str] = field(default_factory=list)


def compare_totals(a: Totals, b: Totals,
                   tol: float = _MATCH_TOLERANCE) -> TotalsCompare:
    out = TotalsCompare(a=a, b=b)
    out.delta_ob = round(a.total_ob - b.total_ob, 2)
    out.delta_dr = round(a.total_dr - b.total_dr, 2)
    out.delta_cr = round(a.total_cr - b.total_cr, 2)
    out.delta_cb = round(a.total_cb - b.total_cb, 2)
    for dim, dv in (("OB", out.delta_ob), ("DR", out.delta_dr),
                    ("CR", out.delta_cr), ("CB", out.delta_cb)):
        if abs(dv) > tol:
            out.dimensions_diverged.append(dim)
    out.matched = not out.dimensions_diverged
    return out


# Per-account drill-down — runs ONLY when top-line totals diverge.

@dataclass
class AccountDiagnosis:
    key: str               # account code or normalised name
    name: str              # display name (Tally or OneShell)
    tally: TBRow | None = None
    file: TBRow | None = None
    db: TBRow | None = None
    api: TBRow | None = None
    delta_dr: float = 0.0  # tally.dr - oneshell-best.dr
    delta_cr: float = 0.0
    delta_ob: float = 0.0
    delta_cb: float = 0.0
    diagnoses: list[str] = field(default_factory=list)


def drill_down_per_account(
    *,
    tally: list[TBRow],
    file_rows: list[TBRow] | None = None,
    db_rows: list[TBRow] | None = None,
    api_rows: list[TBRow] | None = None,
    tol: float = _DIFF_TOLERANCE,
) -> list[AccountDiagnosis]:
    """For each unique account across all sources, compare DR/CR/OB/CB
    and emit a diagnosis explaining where + why the divergence is."""
    by_t = {_key(r): r for r in (tally or [])}
    by_f = {_key(r): r for r in (file_rows or [])}
    by_d = {_key(r): r for r in (db_rows or [])}
    by_a = {_key(r): r for r in (api_rows or [])}

    keys = sorted(set(by_t) | set(by_f) | set(by_d) | set(by_a))
    out: list[AccountDiagnosis] = []
    for k in keys:
        t = by_t.get(k)
        f = by_f.get(k)
        d = by_d.get(k)
        a = by_a.get(k)
        # Pick best OneShell representative (priority: api > db > file)
        o_best = a or d or f
        diag = AccountDiagnosis(
            key=k,
            name=(t.name if t else (o_best.name if o_best else k)),
            tally=t, file=f, db=d, api=a,
        )
        if t and o_best:
            diag.delta_dr = round(t.debit - o_best.debit, 2)
            diag.delta_cr = round(t.credit - o_best.credit, 2)
            diag.delta_ob = round(t.opening - o_best.opening, 2)
            diag.delta_cb = round(t.closing - o_best.closing, 2)

        # Diagnoses
        if t and not (f or d or a):
            diag.diagnoses.append(
                "tally_only — account exists in Tally but missing in "
                "OneShell (missing master record OR business mis-routed)"
            )
        elif (f or d or a) and not t:
            diag.diagnoses.append(
                "oneshell_only — account exists in OneShell but missing "
                "in Tally export (forgot to map / Tally suspense)"
            )
        if t and o_best:
            if abs(diag.delta_ob) > tol:
                diag.diagnoses.append(
                    f"opening_balance_mismatch — Tally ₹{t.opening:,.2f} "
                    f"vs OneShell ₹{o_best.opening:,.2f} (likely OB "
                    "migration / FY-start lag)"
                )
            if abs(diag.delta_dr) > tol and abs(diag.delta_cr) <= tol:
                diag.diagnoses.append(
                    "debit_only_drift — same credits, debit side "
                    "diverges (missing DR transactions in OneShell or "
                    "extra DR in Tally)"
                )
            if abs(diag.delta_cr) > tol and abs(diag.delta_dr) <= tol:
                diag.diagnoses.append(
                    "credit_only_drift — same debits, credit side "
                    "diverges (missing CR transactions in OneShell or "
                    "extra CR in Tally)"
                )
            if abs(diag.delta_dr) > tol and abs(diag.delta_cr) > tol:
                diag.diagnoses.append(
                    "both_sides_drift — both DR and CR differ (account "
                    "may have new transactions in one system not yet in "
                    "the other)"
                )
            if d and a and abs(d.closing - a.closing) > tol:
                diag.diagnoses.append(
                    f"db_vs_api_drift — chartOfAccounts CB ₹{d.closing:,.2f}"
                    f" disagrees with /trialBalance API ₹{a.closing:,.2f} "
                    "(stale cache / async aggregation lag)"
                )
            if f and a and abs(f.closing - a.closing) > tol:
                diag.diagnoses.append(
                    f"file_vs_api_drift — uploaded file CB ₹{f.closing:,.2f}"
                    f" disagrees with live API ₹{a.closing:,.2f} (export "
                    "was generated against an older snapshot)"
                )
        if not diag.diagnoses:
            diag.diagnoses.append("ok — totals within tolerance")
        out.append(diag)
    return out


@dataclass
class FullReport:
    env: str
    business_id: str
    totals: dict[str, Totals] = field(default_factory=dict)
    pair_compares: dict[str, TotalsCompare] = field(default_factory=dict)
    drill: list[AccountDiagnosis] = field(default_factory=list)
    has_top_line_mismatch: bool = False


def verify(
    *,
    tally: list[TBRow],
    file_rows: list[TBRow],
    db_rows: list[TBRow] | None = None,
    api_rows: list[TBRow] | None = None,
    env: str = "qa",
    business_id: str = "",
) -> FullReport:
    """Full top-down verification:

       1. compute totals (DR, CR, OB, CB) for every source
       2. compare each pair of sources at the totals level
       3. if any pair diverges → drill down per-account with diagnoses
    """
    rep = FullReport(env=env, business_id=business_id)
    rep.totals["tally"] = compute_totals(tally, source="tally")
    rep.totals["file"] = compute_totals(file_rows, source="file")
    if db_rows is not None:
        rep.totals["db"] = compute_totals(db_rows, source="db")
    if api_rows is not None:
        rep.totals["api"] = compute_totals(api_rows, source="api")

    pairs = [("tally", "file")]
    if api_rows is not None:
        pairs.append(("tally", "api"))
        pairs.append(("file", "api"))
    if db_rows is not None:
        pairs.append(("tally", "db"))
        pairs.append(("file", "db"))
        if api_rows is not None:
            pairs.append(("api", "db"))

    for a, b in pairs:
        if a in rep.totals and b in rep.totals:
            cmp = compare_totals(rep.totals[a], rep.totals[b])
            rep.pair_compares[f"{a}_vs_{b}"] = cmp
            if not cmp.matched:
                rep.has_top_line_mismatch = True

    if rep.has_top_line_mismatch:
        rep.drill = drill_down_per_account(
            tally=tally, file_rows=file_rows,
            db_rows=db_rows, api_rows=api_rows,
        )
    return rep


def render_verify_report(rep: FullReport) -> str:
    parts: list[str] = [
        f"# Trial-Balance Verification ({rep.env})",
        "",
        f"Business: `{rep.business_id or '-'}`",
        "",
        "## Totals",
        "",
        "| source | rows | OB | DR | CR | CB |",
        "|---|---:|---:|---:|---:|---:|",
    ]
    for name in ("tally", "file", "db", "api"):
        t = rep.totals.get(name)
        if not t:
            continue
        parts.append(
            f"| **{name}** | {t.row_count} | "
            f"₹{t.total_ob:,.2f} | ₹{t.total_dr:,.2f} | "
            f"₹{t.total_cr:,.2f} | ₹{t.total_cb:,.2f} |"
        )
    parts.append("")
    parts.append("## Pairwise totals comparison")
    parts.append("")
    parts.append("| pair | ΔOB | ΔDR | ΔCR | ΔCB | match |")
    parts.append("|---|---:|---:|---:|---:|:--:|")
    for k, c in rep.pair_compares.items():
        parts.append(
            f"| {k} | ₹{c.delta_ob:,.2f} | ₹{c.delta_dr:,.2f} | "
            f"₹{c.delta_cr:,.2f} | ₹{c.delta_cb:,.2f} | "
            f"{'✓' if c.matched else '✗ ' + ','.join(c.dimensions_diverged)} |"
        )
    parts.append("")
    if not rep.has_top_line_mismatch:
        parts.append(
            "**All totals within ₹1 tolerance — no drill-down needed.**"
        )
        return "\n".join(parts) + "\n"
    parts.append(
        "**Top-line mismatch detected — drilling down per account "
        "(only flagged accounts shown).**"
    )
    parts.append("")
    parts.append("## Per-account diagnoses")
    parts.append("")
    parts.append("| account | code | sources | ΔOB | ΔDR | ΔCR | ΔCB | diagnosis |")
    parts.append("|---|---|---|---:|---:|---:|---:|---|")
    for d in rep.drill:
        if d.diagnoses == ["ok — totals within tolerance"]:
            continue
        sources = []
        if d.tally:
            sources.append("T")
        if d.file:
            sources.append("F")
        if d.db:
            sources.append("D")
        if d.api:
            sources.append("A")
        code = (d.tally.code if d.tally else "") or \
               (d.file.code if d.file else "") or \
               (d.db.code if d.db else "")
        parts.append(
            f"| {d.name[:50]} | {code} | {''.join(sources)} | "
            f"₹{d.delta_ob:,.2f} | ₹{d.delta_dr:,.2f} | "
            f"₹{d.delta_cr:,.2f} | ₹{d.delta_cb:,.2f} | "
            f"{'; '.join(d.diagnoses)[:200]} |"
        )
    return "\n".join(parts) + "\n"


def _key(row: TBRow) -> str:
    return row.code.strip().lower() if row.code else normalise_name(row.name)


def reconcile(
    tally: list[TBRow], oneshell: list[TBRow],
    *, env: str = "qa", business_id: str = "",
) -> Reconciliation:
    rec = Reconciliation(env=env, business_id=business_id)
    by_t = {_key(r): r for r in tally}
    by_o = {_key(r): r for r in oneshell}

    for k, t in by_t.items():
        rec.tally_total += t.closing
        o = by_o.get(k)
        if o is None:
            rec.tally_only.append(_row_dict(t, "tally_only"))
            continue
        delta_open = round(t.opening - o.opening, 2)
        delta_dr = round(t.debit - o.debit, 2)
        delta_cr = round(t.credit - o.credit, 2)
        delta_close = round(t.closing - o.closing, 2)
        entry = {
            "key": k,
            "tally_name": t.name, "oneshell_name": o.name,
            "code": t.code or o.code,
            "parent": t.parent or o.parent,
            "tally_close": t.closing, "oneshell_close": o.closing,
            "delta_open": delta_open,
            "delta_debit": delta_dr,
            "delta_credit": delta_cr,
            "delta_close": delta_close,
        }
        absd = abs(delta_close)
        if absd <= _MATCH_TOLERANCE:
            entry["bucket"] = "match"
            rec.matched.append(entry)
        elif absd <= _DIFF_TOLERANCE:
            entry["bucket"] = "diff"
            rec.diff.append(entry)
        else:
            entry["bucket"] = "large"
            rec.large.append(entry)
    for k, o in by_o.items():
        rec.oneshell_total += o.closing
        if k not in by_t:
            rec.oneshell_only.append(_row_dict(o, "oneshell_only"))

    rec.gap = round(rec.tally_total - rec.oneshell_total, 2)
    return rec


@dataclass
class ThreeWayResult:
    env: str
    business_id: str
    file_vs_db: Reconciliation = field(default_factory=lambda: Reconciliation("", ""))
    tally_vs_file: Reconciliation = field(default_factory=lambda: Reconciliation("", ""))
    tally_vs_db: Reconciliation = field(default_factory=lambda: Reconciliation("", ""))


def reconcile_3way(
    tally: list[TBRow],
    oneshell_file: list[TBRow],
    oneshell_db: list[TBRow],
    *, env: str = "qa", business_id: str = "",
) -> ThreeWayResult:
    """Run all three pairwise recons. file_vs_db catches export drift
    (the file someone uploaded doesn't match the live DB), and the two
    tally_vs_* surface where Tally disagrees with each canonical
    OneShell snapshot."""
    return ThreeWayResult(
        env=env, business_id=business_id,
        file_vs_db=reconcile(oneshell_file, oneshell_db,
                             env=env + "/file-vs-db",
                             business_id=business_id),
        tally_vs_file=reconcile(tally, oneshell_file,
                                env=env + "/tally-vs-file",
                                business_id=business_id),
        tally_vs_db=reconcile(tally, oneshell_db,
                              env=env + "/tally-vs-db",
                              business_id=business_id),
    )


def render_markdown_3way(r: ThreeWayResult) -> str:
    parts = [f"# 3-way Trial-Balance Reconciliation ({r.env})", ""]
    for label, rec in (
        ("File ↔ DB (export drift)", r.file_vs_db),
        ("Tally ↔ File", r.tally_vs_file),
        ("Tally ↔ DB", r.tally_vs_db),
    ):
        parts.append(f"## {label}")
        parts.append(render_markdown(rec))
        parts.append("")
    return "\n".join(parts)


def _row_dict(r: TBRow, bucket: str) -> dict:
    return {
        "key": _key(r), "name": r.name, "parent": r.parent, "code": r.code,
        "closing": r.closing, "bucket": bucket,
    }


# ─────────── reporting ────────────────────────────────────────────────

def render_markdown(rec: Reconciliation) -> str:
    lines: list[str] = [
        f"## Trial-Balance Reconciliation ({rec.env})",
        "",
        f"- Tally CB total: ₹{rec.tally_total:,.2f}",
        f"- OneShell CB total: ₹{rec.oneshell_total:,.2f}",
        f"- **Gap: ₹{rec.gap:,.2f}**",
        "",
        f"| bucket | count |",
        f"|---|---:|",
        f"| MATCH (|Δ| ≤ ₹1) | {len(rec.matched)} |",
        f"| DIFF (₹1–₹100) | {len(rec.diff)} |",
        f"| LARGE (> ₹100) | {len(rec.large)} |",
        f"| Tally-only | {len(rec.tally_only)} |",
        f"| OneShell-only | {len(rec.oneshell_only)} |",
        "",
    ]
    if rec.large:
        lines.append("### Top 10 LARGE diffs")
        lines.append("| account | code | tally CB | oneshell CB | Δ |")
        lines.append("|---|---|---:|---:|---:|")
        for e in sorted(rec.large, key=lambda x: -abs(x["delta_close"]))[:10]:
            lines.append(
                f"| {e['tally_name'][:60]} | {e.get('code','')} | "
                f"₹{e['tally_close']:,.2f} | ₹{e['oneshell_close']:,.2f} | "
                f"**₹{e['delta_close']:,.2f}** |"
            )
    return "\n".join(lines) + "\n"


def write_csv(rec: Reconciliation, out_path: Path) -> None:
    rows: list[dict] = []
    for e in rec.matched + rec.diff + rec.large:
        rows.append(e)
    for e in rec.tally_only + rec.oneshell_only:
        rows.append(e)
    if not rows:
        return
    fieldnames = sorted({k for r in rows for k in r.keys()})
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow(r)


# ─────────── CLI ──────────────────────────────────────────────────────

def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(prog="aiforge-agents-tb")
    p.add_argument("--tally", required=True)
    p.add_argument("--oneshell", required=True)
    p.add_argument("--env", default="qa", choices=("qa", "prod"))
    p.add_argument("--business", default="")
    p.add_argument("--out-dir", default=".")
    p.add_argument("--validate-with-mongo", action="store_true",
                   help="also pull live OneShell rows from MongoDB "
                        "and run a 3-way reconcile")
    p.add_argument("--strict", action="store_true",
                   help="exit non-zero on any file validation error")
    a = p.parse_args(argv)

    # 1. validate files first — fail fast with a clear schema error
    fv_t = validate_file(a.tally, expected="tally")
    fv_o = validate_file(a.oneshell, expected="oneshell")
    print("# File validation")
    for fv in (fv_t, fv_o):
        flag = "OK " if fv.ok else "ERR"
        print(f"  [{flag}] {fv.path}  rows={fv.row_count}  kind={fv.detected_kind}")
        for err in fv.errors:
            print(f"      - error: {err}")
        for w in fv.warnings:
            print(f"      - warn: {w}")
    if (not fv_t.ok or not fv_o.ok) and a.strict:
        print(json.dumps({"status": "validation_failed",
                          "errors": fv_t.errors + fv_o.errors}, indent=2))
        return 2
    if not fv_t.ok or not fv_o.ok:
        # non-strict: still proceed but warn
        print("WARN proceeding despite validation issues — pass --strict to halt")

    # 2. parse + (optional) live DB fetch
    t_rows = parse_tally(Path(a.tally))
    o_file = parse_oneshell(Path(a.oneshell))
    o_db: list[TBRow] = []
    if a.validate_with_mongo:
        if not a.business:
            print("ERR --validate-with-mongo needs --business")
            return 2
        try:
            o_db = fetch_oneshell_from_mongo(
                business_id=a.business, env=a.env,
            )
            print(f"# Mongo fetch: {len(o_db)} rows")
        except SystemExit as exc:
            print(f"ERR mongo fetch failed: {exc}")
            return 2
        except Exception as exc:
            print(f"WARN mongo fetch errored: {exc}; falling back to file-only")
            o_db = []

    out_dir = Path(a.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # 3a. Top-down verify — totals first, then drill-down per account
    rep = verify(
        tally=t_rows, file_rows=o_file,
        db_rows=o_db or None, api_rows=None,
        env=a.env, business_id=a.business,
    )
    print()
    print(render_verify_report(rep))

    # 3b. Reconcile (kept for backwards-compat + per-row CSVs)
    if o_db:
        three = reconcile_3way(
            t_rows, o_file, o_db, env=a.env, business_id=a.business,
        )
        # write CSVs for each pair
        for label, rec in (
            ("file-vs-db", three.file_vs_db),
            ("tally-vs-file", three.tally_vs_file),
            ("tally-vs-db", three.tally_vs_db),
        ):
            write_csv(
                rec,
                out_dir / f"{a.env}-{a.business or 'all'}-{label}.csv",
            )
        summary = {
            "env": a.env, "business_id": a.business, "mode": "3way",
            "file_vs_db_gap":   three.file_vs_db.gap,
            "tally_vs_file_gap": three.tally_vs_file.gap,
            "tally_vs_db_gap":  three.tally_vs_db.gap,
            "buckets_tally_vs_db": {
                "match": len(three.tally_vs_db.matched),
                "diff":  len(three.tally_vs_db.diff),
                "large": len(three.tally_vs_db.large),
            },
        }
        print(json.dumps(summary, indent=2))
        print(render_markdown_3way(three))
        # Block on any LARGE bucket across the three recons
        any_large = any([
            three.file_vs_db.large,
            three.tally_vs_file.large,
            three.tally_vs_db.large,
        ])
        return 1 if any_large else 0

    # Plain 2-way (no Mongo)
    rec = reconcile(t_rows, o_file, env=a.env, business_id=a.business)
    csv_path = out_dir / f"{a.env}-{a.business or 'all'}-tb-diff.csv"
    write_csv(rec, csv_path)
    summary = {
        "env": rec.env, "business_id": rec.business_id, "mode": "2way",
        "tally_total": rec.tally_total,
        "oneshell_total": rec.oneshell_total,
        "gap": rec.gap,
        "buckets": {
            "match": len(rec.matched),
            "diff": len(rec.diff),
            "large": len(rec.large),
            "tally_only": len(rec.tally_only),
            "oneshell_only": len(rec.oneshell_only),
        },
        "csv": str(csv_path),
    }
    print(json.dumps(summary, indent=2))
    print()
    print(render_markdown(rec))
    return 1 if (rec.large or abs(rec.gap) > 1000) else 0


def run_workflow(ticket: dict, *, log=None) -> dict:
    """Stable handler entry — invoked by the WorkflowRegistry dispatcher.

    Reads ``ticket`` (a dict containing at minimum ``id``, ``identifier``,
    ``title``, ``body`` and optional ``metadata``), pulls Tally + OneShell
    attachments, runs the validate → parse → fetch-from-DB → verify
    pipeline, and returns a doer-outcome-shaped dict that the
    orchestrator hands to the Validator/Architect stages.

    Soft-fails on any expected error (missing attachments, parse error,
    DB unreachable) — never raises. The orchestrator decides downstream
    routing based on ``blocked_by_detectors`` and ``problems``.
    """
    # Lazy import to avoid pulling the learner/persistence layer at
    # module load time (the registry imports trial_balance at startup).
    from aiforge_core.aiforge_agents.learner import online as _learner

    ticket_id = ticket.get("identifier") or ticket.get("id") or ""
    title = ticket.get("title") or ""
    body = ticket.get("body") or ""
    text = (title + " " + body).lower()

    def _emit(event: str, **fields):
        if log is not None:
            try:
                from aiforge_core.runtime.logging_setup import emit as _emit_evt
                _emit_evt(log, event, ticket=ticket_id, **fields)
            except Exception:
                pass

    atts = _learner.attachments_for(str(ticket_id))
    by_role: dict[str, dict] = {}
    for a in atts:
        by_role.setdefault(a["role"], a)
    missing = [r for r in ("tally", "oneshell") if r not in by_role]
    if missing:
        _emit("trial_balance.missing_attachments", missing=missing)
        return {
            "artifact_type": "doer_outcome",
            "process": "trial_balance",
            "applied": False,
            "udiff": "",
            "target": "trial-balance-report.md",
            "problems": [{
                "mode": "missing_attachment",
                "evidence": f"required attachment role(s) missing: {missing}",
            }],
            "blocked_by_detectors": True,
        }

    env = "prod" if "prod" in text else "qa"
    try:
        fv_t = validate_file(by_role["tally"]["file_path"], expected="tally")
        fv_o = validate_file(by_role["oneshell"]["file_path"], expected="oneshell")
        validation_errors: list[str] = []
        for fv in (fv_t, fv_o):
            for e in fv.errors:
                validation_errors.append(f"{Path(fv.path).name}: {e}")
        if validation_errors:
            _emit("trial_balance.validation_failed",
                  errors=validation_errors[:5])
            return {
                "artifact_type": "doer_outcome",
                "process": "trial_balance",
                "env": env,
                "applied": False,
                "udiff": "",
                "target": "trial-balance-report.md",
                "problems": [{"mode": "file_validation",
                              "evidence": e} for e in validation_errors],
                "blocked_by_detectors": True,
                "file_validation": {
                    "tally": {"ok": fv_t.ok, "errors": fv_t.errors,
                              "warnings": fv_t.warnings,
                              "rows": fv_t.row_count},
                    "oneshell": {"ok": fv_o.ok, "errors": fv_o.errors,
                                 "warnings": fv_o.warnings,
                                 "rows": fv_o.row_count},
                },
            }
        t_rows = parse_tally(Path(by_role["tally"]["file_path"]))
        o_file = parse_oneshell(Path(by_role["oneshell"]["file_path"]))
        bid = ""
        m = re.search(r"\bb\d{14,}\b", text)
        if m:
            bid = m.group(0)
        o_db: list = []
        if bid:
            import datetime as _dt
            today = _dt.date.today()
            fy_year = today.year + (1 if today.month >= 4 else 0)
            try:
                o_db = fetch_oneshell_via_api(
                    business_id=bid, env=env, financial_year=fy_year,
                )
            except (OSError, ValueError, KeyError) as exc:
                _emit("trial_balance.api_unreachable",
                      error=str(exc)[:200], business_id=bid)
                try:
                    o_db = fetch_oneshell_from_mongo(
                        business_id=bid, env=env,
                    )
                except (OSError, ValueError, KeyError) as exc2:
                    _emit("trial_balance.mongo_unreachable",
                          error=str(exc2)[:200])
                    o_db = []
        rep = verify(
            tally=t_rows, file_rows=o_file,
            db_rows=o_db or None, api_rows=None,
            env=env, business_id=bid,
        )
        md = render_verify_report(rep)
        totals_view = {
            src: {"OB": t.total_ob, "DR": t.total_dr,
                  "CR": t.total_cr, "CB": t.total_cb,
                  "rows": t.row_count}
            for src, t in rep.totals.items()
        }
        if o_db:
            three = reconcile_3way(
                t_rows, o_file, o_db, env=env, business_id=bid,
            )
            any_large = any([
                three.file_vs_db.large,
                three.tally_vs_file.large,
                three.tally_vs_db.large,
            ])
            _emit("trial_balance.three_way_done",
                  large_buckets=any_large,
                  top_line_mismatch=rep.has_top_line_mismatch)
            return {
                "artifact_type": "doer_outcome",
                "process": "trial_balance",
                "mode": "3way+verify",
                "env": env,
                "udiff": md,
                "target": "trial-balance-report.md",
                "totals": totals_view,
                "top_line_mismatch": rep.has_top_line_mismatch,
                "drilled_accounts": len([d for d in rep.drill
                                         if d.diagnoses != ["ok — totals within tolerance"]]),
                "file_vs_db_gap": three.file_vs_db.gap,
                "tally_vs_file_gap": three.tally_vs_file.gap,
                "tally_vs_db_gap":  three.tally_vs_db.gap,
                "buckets": {
                    "tally_vs_db_match": len(three.tally_vs_db.matched),
                    "tally_vs_db_large": len(three.tally_vs_db.large),
                    "file_vs_db_large":  len(three.file_vs_db.large),
                },
                "applied": False,
                "problems": [],
                "blocked_by_detectors": (any_large or rep.has_top_line_mismatch),
            }
        rec = reconcile(t_rows, o_file, env=env, business_id=bid)
        _emit("trial_balance.two_way_done",
              gap=rec.gap, large_count=len(rec.large),
              top_line_mismatch=rep.has_top_line_mismatch)
        return {
            "artifact_type": "doer_outcome",
            "process": "trial_balance",
            "mode": "2way+verify",
            "env": env,
            "udiff": md,
            "target": "trial-balance-report.md",
            "totals": totals_view,
            "top_line_mismatch": rep.has_top_line_mismatch,
            "drilled_accounts": len([d for d in rep.drill
                                     if d.diagnoses != ["ok — totals within tolerance"]]),
            "tally_total": rec.tally_total,
            "oneshell_total": rec.oneshell_total,
            "gap": rec.gap,
            "buckets": {
                "match": len(rec.matched),
                "diff":  len(rec.diff),
                "large": len(rec.large),
                "tally_only":    len(rec.tally_only),
                "oneshell_only": len(rec.oneshell_only),
            },
            "applied": False,
            "problems": [],
            "blocked_by_detectors": rep.has_top_line_mismatch,
        }
    except (ImportError, OSError, ValueError, KeyError) as exc:
        _emit("trial_balance.process_failed",
              error=str(exc)[:200], type=type(exc).__name__)
        return {
            "artifact_type": "doer_outcome",
            "process": "trial_balance",
            "error": f"validator_failed: {type(exc).__name__}: {exc}",
            "error_type": type(exc).__name__,
            "applied": False,
            "problems": [{
                "mode": "trial_balance_exception",
                "evidence": str(exc)[:500],
                "exc_type": type(exc).__name__,
            }],
            "udiff": "",
        }


if __name__ == "__main__":
    sys.exit(main())

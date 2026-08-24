"""Document → text extraction, page-aware.

Owns everything that turns a pdf / docx / xlsx / text file into readable text:
per-page text, structured tables, page segmentation, and page-range selection.
Vision-dependent bits (image captions, scanned-page OCR) live in ``chat_media``
— this module is pure text and has no LLM/vision dependency.

Pages:
  • PDF   — real pages (pypdf), one entry per page.
  • docx  — Word has no page model in the file; we segment on rendered page
            breaks (``w:lastRenderedPageBreak``, written by Word) and manual
            page breaks (``w:br w:type="page"``). A docx never opened in Word
            carries no breaks → we fall back to fixed-size character "pages" so
            page selection still works (approximate, flagged).
  • xlsx  — each worksheet is treated as a "page".
  • text  — one page, or char-approx pages when large.

Everything is bounded (page cap + char budget) so a 400-page document folds
without blowing memory. Best-effort — every entry point swallows errors and
returns "" / [] rather than raising, so an upload never breaks.
"""
from __future__ import annotations

import os
from pathlib import Path

_W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
_DRAWING = _W + "drawing"
_PICT = _W + "pict"
_LASTRENDERED = _W + "lastRenderedPageBreak"
_BR = _W + "br"
_BR_TYPE = _W + "type"

_EXT_MIME = {
    ".pdf": "application/pdf",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".xls": "application/vnd.ms-excel",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".csv": "text/csv", ".txt": "text/plain", ".md": "text/markdown",
    ".json": "application/json", ".log": "text/plain", ".yaml": "text/yaml",
    ".yml": "text/yaml", ".py": "text/x-python", ".js": "text/javascript",
    ".ts": "text/plain", ".java": "text/x-java", ".go": "text/x-go",
}
_TEXT_EXTS = {".txt", ".md", ".csv", ".json", ".log", ".yaml", ".yml",
              ".py", ".js", ".ts", ".java", ".go", ".sh", ".sql", ".html",
              ".xml", ".toml", ".ini", ".cfg"}


def _int_env(name: str, default: int) -> int:
    try:
        return int(os.environ.get(name, "") or default)
    except ValueError:
        return default


# A 100+-page document must extract fully, but page-by-page and bounded so a
# giant file can't blow memory or the model's context. Both env-tunable.
def _pdf_page_cap() -> int:
    return _int_env("AIFORGE_DOC_MAX_PAGES", 500)


def _doc_char_budget() -> int:
    # ~2k chars/page → 1M chars ≈ 500 pages of extractable text.
    return _int_env("AIFORGE_DOC_MAX_CHARS", 1_000_000)


def _approx_page_chars() -> int:
    # ~one printed page of prose ≈ 3000 chars (≈500 words). Used to estimate
    # docx/text page numbers when the file carries no reliable page breaks.
    return _int_env("AIFORGE_APPROX_PAGE_CHARS", 3000)


def _is_pdf(path: str, mime: str) -> bool:
    return os.path.splitext(path)[1].lower() == ".pdf" \
        or (mime or "") == "application/pdf"


def _is_docx(path: str, mime: str) -> bool:
    return os.path.splitext(path)[1].lower() == ".docx" \
        or "wordprocessing" in (mime or "")


# ── PDF ──────────────────────────────────────────────────────────────────
def _pdf_tables(path: str, page_cap: int) -> dict[int, list[str]]:
    """Structured tables per page via pdfplumber (optional dep). Returns
    ``{page_index: ["a | b", …]}``. Empty when pdfplumber is absent or the PDF
    has no detectable tables — pypdf's flat text still carries the numbers, this
    just restores the grid. Best-effort, never raises."""
    try:
        import pdfplumber
    except Exception:  # noqa: BLE001 — optional; flat text is the fallback
        return {}
    tables: dict[int, list[str]] = {}
    try:
        with pdfplumber.open(path) as pdf:
            for i, page in enumerate(pdf.pages[:page_cap]):
                rows: list[str] = []
                for tbl in (page.extract_tables() or []):
                    for row in tbl:
                        cells = ["" if c is None else str(c).replace("\n", " ")
                                 for c in row]
                        rows.append(" | ".join(cells))
                if rows:
                    tables[i] = rows
    except Exception:  # noqa: BLE001
        return tables
    return tables


def _pdf_pages(path: str) -> list[str]:
    """One text entry per PDF page (extracted text + interleaved tables),
    bounded by the page cap."""
    from pypdf import PdfReader
    r = PdfReader(path)
    cap = _pdf_page_cap()
    tables = _pdf_tables(path, cap)
    pages: list[str] = []
    for i, p in enumerate(r.pages[:cap]):
        try:
            t = p.extract_text() or ""
        except Exception:  # noqa: BLE001
            t = ""
        if i in tables:
            t = (t + "\n[tables]\n" + "\n".join(tables[i])).strip()
        pages.append(t)
    return pages


# ── docx ─────────────────────────────────────────────────────────────────
def _para_has_page_break(p_elem) -> bool:
    """A paragraph that renders (or forces) a page break — Word's auto
    ``lastRenderedPageBreak`` or a manual ``<w:br w:type="page"/>``."""
    if p_elem.findall(".//" + _LASTRENDERED):
        return True
    for br in p_elem.findall(".//" + _BR):
        if br.get(_BR_TYPE) == "page":
            return True
    return False


def _docx_pages(path: str) -> list[str]:
    """docx text segmented into pages on rendered/manual page breaks; each page
    carries paragraphs AND tables (in document order) plus ``[embedded image N]``
    markers. One page when the doc has no break info (never opened in Word)."""
    import docx
    from docx.oxml.table import CT_Tbl
    from docx.oxml.text.paragraph import CT_P
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    doc = docx.Document(path)
    pages: list[list[str]] = [[]]
    img_n = 0
    for child in doc.element.body.iterchildren():
        if isinstance(child, CT_P):
            if _para_has_page_break(child):
                pages.append([])                 # break starts a new page
            para = Paragraph(child, doc)
            if para.text.strip():
                pages[-1].append(para.text)
            if child.findall(".//" + _DRAWING) or child.findall(".//" + _PICT):
                img_n += 1
                pages[-1].append(f"[embedded image {img_n}]")
        elif isinstance(child, CT_Tbl):
            table = Table(child, doc)
            for row in table.rows:
                cells = [c.text.strip().replace("\n", " ") for c in row.cells]
                pages[-1].append(" | ".join(cells))
    return ["\n".join(p).strip() for p in pages]


# ── xlsx ─────────────────────────────────────────────────────────────────
def _xlsx_pages(path: str) -> list[str]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    pages: list[str] = []
    for ws in wb.worksheets:
        out = [f"# sheet: {ws.title}"]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 500:
                out.append("… (more rows)")
                break
            out.append(", ".join("" if c is None else str(c) for c in row))
        pages.append("\n".join(out))
    return pages


# ── page model / char approximation ────────────────────────────────────────
def _approx_pages(text: str, size: int) -> list[str]:
    """Split a page-less document into fixed-size character "pages" (on line
    boundaries) so page selection still works. Approximate — flagged by the
    caller."""
    lines = text.splitlines()
    pages: list[str] = []
    buf: list[str] = []
    used = 0
    for ln in lines:
        if used + len(ln) > size and buf:
            pages.append("\n".join(buf))
            buf, used = [], 0
        buf.append(ln)
        used += len(ln) + 1
    if buf:
        pages.append("\n".join(buf))
    return pages or [text]


def _docx_reported_pages(path: str) -> int | None:
    """Word's OWN page count, read from ``docProps/app.xml`` inside the docx zip
    (the ``<Pages>`` element Word writes on save). This is the authoritative
    total — no renderer needed. None when absent (e.g. a docx never saved by
    Word) or unreadable. stdlib only (zipfile + ElementTree)."""
    import xml.etree.ElementTree as ET
    import zipfile
    try:
        with zipfile.ZipFile(path) as z:
            with z.open("docProps/app.xml") as f:
                root = ET.parse(f).getroot()
    except Exception:  # noqa: BLE001 — no app.xml / not a zip / malformed
        return None
    for el in root.iter():
        if el.tag.rsplit("}", 1)[-1] == "Pages":
            try:
                n = int((el.text or "").strip())
                return n if n > 0 else None
            except (ValueError, TypeError):
                return None
    return None


def _split_into_n(text: str, n: int) -> list[str]:
    """Split text into EXACTLY ``n`` char-proportional pages, to honour Word's
    own page count. Each line is assigned to a page by its character position
    (``pos/total * n``) so the page count is precise regardless of line length;
    with far fewer lines than pages we fall back to a raw char slice so every
    page still carries content."""
    if n <= 1:
        return [text]
    lines = text.splitlines()
    total = max(1, len(text))
    if len(lines) < n:                       # too few lines to fill n pages
        size = total / n
        return [text[round(i * size):round((i + 1) * size)] for i in range(n)]
    buckets: list[list[str]] = [[] for _ in range(n)]
    pos = 0
    for ln in lines:
        idx = min(n - 1, int(pos * n / total))
        buckets[idx].append(ln)
        pos += len(ln) + 1
    return ["\n".join(b) for b in buckets]


def _docx_paginate(path: str) -> tuple[list[str], str]:
    """docx → (pages, kind). Page count, best source first:

      1. Word's own ``<Pages>`` from docProps/app.xml — authoritative TOTAL;
         text is mapped to that many pages proportionally (kind "word": count is
         exact, per-page content is an approximate slice).
      2. Rendered page breaks (``lastRenderedPageBreak``) when DENSE enough to
         be real pages — Word writes them sporadically, so a 300-page report can
         carry ~10 markers; trusting sparse markers under-counts massively and
         makes a high page number fall out of range (kind "exact").
      3. Char-approx pagination spanning the whole doc (kind "approx").
    """
    import math
    breakpages = _docx_pages(path)
    whole = "\n".join(breakpages).strip()
    if not whole:
        return [], "none"
    reported = _docx_reported_pages(path)
    if reported and reported > 1:
        return _split_into_n(whole, reported), "word"
    nonempty = [p for p in breakpages if p.strip()]
    approx_n = max(1, math.ceil(len(whole) / _approx_page_chars()))
    if len(nonempty) >= 2 and len(breakpages) >= approx_n * 0.6:
        return breakpages, "exact"
    if len(whole) > _approx_page_chars():
        return _approx_pages(whole, _approx_page_chars()), "approx"
    return [whole], "approx"


def paginate(path: str, mime: str = "") -> tuple[list[str], str]:
    """(pages, kind) for a document. kind ∈ {exact, approx, sheets, none}.
    PDF → real pages ("exact"); docx → break-segmented when dense enough else
    char-approx; xlsx → per-sheet ("sheets"); text → char-approx when large.
    Never raises → ([], "none") on any failure."""
    ext = os.path.splitext(path)[1].lower()
    try:
        if _is_pdf(path, mime):
            return _pdf_pages(path), "exact"
        if ext == ".xlsx" or "spreadsheet" in (mime or ""):
            return _xlsx_pages(path), "sheets"
        if _is_docx(path, mime):
            return _docx_paginate(path)
        if (mime or "").startswith("text/") or ext in _TEXT_EXTS:
            txt = Path(path).read_text(encoding="utf-8", errors="ignore")
            if len(txt) > _approx_page_chars():
                return _approx_pages(txt, _approx_page_chars()), "approx"
            return ([txt], "exact") if txt else ([], "none")
    except Exception:  # noqa: BLE001
        return [], "none"
    return [], "none"


def document_pages(path: str, mime: str = "") -> list[str]:
    """The document as a list of page texts (page 1 = index 0). [] when nothing
    readable. Never raises. See :func:`paginate` for the exact/approx kind."""
    return paginate(path, mime)[0]


def pagination_kind(path: str, mime: str = "") -> str:
    """How page numbers were derived: "exact" (PDF pages / dense docx breaks),
    "word" (docx total is Word's own count; per-page content mapped
    proportionally), "approx" (char-estimated), "sheets" (xlsx), or "none"."""
    return paginate(path, mime)[1]


def _join_pages(pages: list[str], budget: int, *,
                numbers: list[int] | None = None) -> str:
    """Join page texts with ``--- page N ---`` markers, stopping at the char
    budget. ``numbers`` overrides the 1-based labels (for a selected subset)."""
    out: list[str] = []
    used = 0
    for i, page in enumerate(pages):
        n = numbers[i] if numbers else i + 1
        block = f"--- page {n} ---\n{page}"
        out.append(block)
        used += len(block)
        if used >= budget:
            out.append(f"… (truncated at page {n}, {budget} char budget reached)")
            break
    return "\n\n".join(out)


def extract_text(path: str, mime: str = "") -> str:
    """Full readable text of a document, with ``--- page N ---`` markers so the
    reader (and the model) knows which page each part came from. Bounded by the
    char budget. Best-effort, never raises."""
    ext = os.path.splitext(path)[1].lower()
    try:
        if _is_pdf(path, mime) or _is_docx(path, mime) \
                or ext == ".xlsx" or "spreadsheet" in (mime or ""):
            pages = document_pages(path, mime)
            return _join_pages(pages, _doc_char_budget())
        if (mime or "").startswith("text/") or ext in _TEXT_EXTS:
            return Path(path).read_text(encoding="utf-8", errors="ignore")
    except Exception:  # noqa: BLE001
        return ""
    return ""


# ── page selection ─────────────────────────────────────────────────────────
def _spec_bounds(part: str) -> tuple[int, int] | None:
    """The inclusive 1-based ``(lo, hi)`` a single spec term covers — ``"7"`` is
    ``(7, 7)``, ``"9-5"`` is normalised to ``(5, 9)``. None when it is not a
    number (or a pair of them), which the caller skips."""
    try:
        if "-" not in part:
            n = int(part)
            return (n, n)
        a, b = part.split("-", 1)
        lo, hi = int(a), int(b)
        return (hi, lo) if lo > hi else (lo, hi)
    except ValueError:
        return None


def parse_page_spec(spec: str, total: int) -> list[int]:
    """Parse a page spec like ``"10-20"``, ``"3,5,7-9"``, ``"12"`` into sorted,
    unique 0-based indices within ``[0, total)``. "" / garbage → []."""
    if not spec or total <= 0:
        return []
    idx: set[int] = set()
    for part in str(spec).replace(" ", "").split(","):
        bounds = _spec_bounds(part) if part else None
        if bounds is None:
            continue
        lo, hi = bounds
        # Clamped rather than rejected: "10-99" on a 12-page file means pages
        # 10..12, which is what a reader asking for "10 onwards" meant.
        idx.update(n - 1 for n in range(max(lo, 1), min(hi, total) + 1))
    return sorted(idx)


def extract_pages(path: str, mime: str, spec: str) -> str:
    """Text of ONLY the requested pages (spec e.g. ``"10-20"``), with page
    markers. "" when the spec selects nothing or nothing is readable."""
    pages = document_pages(path, mime)
    idx = parse_page_spec(spec, len(pages))
    if not idx:
        return ""
    selected = [pages[i] for i in idx]
    numbers = [i + 1 for i in idx]
    return _join_pages(selected, _doc_char_budget(), numbers=numbers)


def page_count(path: str, mime: str = "") -> int:
    """How many pages the document has (as this module segments them)."""
    return len(document_pages(path, mime))


__all__ = ["extract_text", "extract_pages", "document_pages", "page_count",
           "parse_page_spec", "paginate", "pagination_kind"]
